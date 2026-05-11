"""Build the week-ahead Excel — 2 columns only: Timestamps and Load."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parents[1]))

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

ROOT = Path(__file__).parents[1]
SUB  = ROOT / "day 2" / "submission 1"
SUB.mkdir(parents=True, exist_ok=True)

# Re-run the prediction (or load existing CSV if present)
import subprocess

# Read the existing CSV produced by predict_week_ahead.py
src_csv = ROOT / "day 2" / "submission 1" / "WeekAhead_Jan1to7_2026.csv"
if not src_csv.exists():
    raise FileNotFoundError(f"Run predict_week_ahead.py first; expected {src_csv}")
df = pd.read_csv(src_csv)
ts_col = "Timestamps" if "Timestamps" in df.columns else "timestamp"
load_col = "Load" if "Load" in df.columns else "load_kw_predicted"
df[ts_col] = pd.to_datetime(df[ts_col])
df = df.rename(columns={ts_col: "timestamp", load_col: "load_kw_predicted"})

# Rename columns per user spec: "Timestamps" and "Load"
df_out = df[["timestamp", "load_kw_predicted"]].rename(
    columns={"timestamp": "Timestamps", "load_kw_predicted": "Load"})
print(f"Forecast rows: {len(df_out)}")

# ── Plot for reference ──────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(15, 4.5))
ax.plot(df_out["Timestamps"], df_out["Load"], color="tab:blue", lw=1.0, label="Forecast")
ax.set_xlabel("Timestamp"); ax.set_ylabel("Load (kW)")
ax.set_title(f"Week-ahead forecast — Jan 1-7, 2026  (8-bag LGBM, long-lag-only features)", fontsize=11)
ax.legend(loc="upper right", fontsize=9)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%a %b %d"))
ax.grid(True, alpha=0.3)
fig.tight_layout()
plot = SUB / "WeekAhead_Jan1to7_2026_plot.png"
fig.savefig(plot, dpi=120, bbox_inches="tight"); plt.close(fig)
print(f"Plot saved -> {plot}")

# ── Excel: simple 2 columns ─────────────────────────────────────────
out_xlsx = SUB / "WeekAhead_Jan1to7_2026.xlsx"
with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    df_out.to_excel(writer, sheet_name="Forecast", index=False)

# Embed the plot
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
wb = load_workbook(out_xlsx)
plots_ws = wb.create_sheet("Plot")
plots_ws["A1"] = "Week-ahead forecast: Jan 1-7, 2026"
img = XLImage(str(plot))
plots_ws.add_image(img, "A3")
wb.save(out_xlsx)
print(f"Excel saved -> {out_xlsx}")

# Also save a clean CSV copy
df_out.to_csv(SUB / "WeekAhead_Jan1to7_2026.csv", index=False)

# Stats
print(f"\nForecast stats:")
print(f"  mean: {df_out['Load'].mean():.3f} kW")
print(f"  std:  {df_out['Load'].std():.3f} kW")
print(f"  min:  {df_out['Load'].min():.3f} kW")
print(f"  max:  {df_out['Load'].max():.3f} kW")
print(f"\nFiles in {SUB}:")
for f in sorted(SUB.iterdir()):
    print(f"  {f.name}  ({f.stat().st_size/1024:.0f} KB)")
