"""
Build the surprise-dataset submission Excel:
- Sheet 'Forecast'  : two columns [timestamp, load_pred]
- Sheet 'Metrics'   : RMSE, MAE, MAPE, sMAPE, NRMSE
- Sheet 'Comparison': all candidate models we tried
- Embedded plots: full-month + first-week zoom + (actual vs forecast)
"""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parents[1]))

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

ROOT = Path(__file__).parents[1]
PLOT_DIR = ROOT / "outputs/plots/surprise"
PLOT_DIR.mkdir(parents=True, exist_ok=True)

# ── Load all artefacts ───────────────────────────────────────────────
df_full = pd.read_parquet(ROOT / "data/features/features_surprise_all.parquet")
df_full["timestamp"] = pd.to_datetime(df_full["timestamp"])
test_mask = (df_full["timestamp"].dt.year == 2026) & (df_full["timestamp"].dt.month == 3)
actuals = df_full[test_mask][["timestamp","load_kw"]].sort_values("timestamp").reset_index(drop=True)

forecast = pd.read_csv(ROOT / "outputs/forecasts/surprise_FINAL_test_preds.csv", parse_dates=["timestamp"])
merged   = actuals.merge(forecast, on="timestamp", how="left")
merged["load_pred"] = merged["load_pred"].ffill().bfill()

ts = merged["timestamp"]
y  = merged["load_kw"].values
p  = merged["load_pred"].values

metrics = json.loads((ROOT / "outputs/models/surprise_metrics.json").read_text())
print(f"Loaded {len(merged)} rows; metrics: {metrics['metrics']}")

# ── Plot 1: full month ────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(15, 4.5))
ax.plot(ts, y, color="black", lw=0.8, alpha=0.85, label="Actual load")
ax.plot(ts, p, color="tab:blue", lw=0.8, alpha=0.85, label="Forecast")
ax.fill_between(ts, y, p, alpha=0.15, color="tab:blue")
ax.set_xlabel("Date"); ax.set_ylabel("Load (kW)")
ax.set_title(f"Surprise dataset — March 2026 forecast vs actual  |  NRMSE = {metrics['metrics']['nrmse']:.2f}%   RMSE = {metrics['metrics']['rmse']:.3f} kW",
             fontsize=11)
ax.legend(loc="upper right", fontsize=9)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
ax.grid(True, alpha=0.3)
fig.tight_layout()
plot_full = PLOT_DIR / "surprise_forecast_full.png"
fig.savefig(plot_full, dpi=120, bbox_inches="tight"); plt.close(fig)

# ── Plot 2: first week zoom ──────────────────────────────────────────
zoom = ts.dt.day <= 7
fig, ax = plt.subplots(figsize=(15, 4.5))
ax.plot(ts[zoom], y[zoom], color="black", lw=1.0, alpha=0.85, label="Actual load")
ax.plot(ts[zoom], p[zoom], color="tab:blue", lw=1.0, alpha=0.85, label="Forecast")
ax.fill_between(ts[zoom], y[zoom], p[zoom], alpha=0.15, color="tab:blue")
ax.set_xlabel("Date"); ax.set_ylabel("Load (kW)")
def nrmse(yy, pp): return float(np.sqrt(np.mean((yy-pp)**2)) / np.mean(yy) * 100)
ax.set_title(f"First week of March 2026 (zoom)  |  NRMSE = {nrmse(y[zoom], p[zoom]):.2f}%", fontsize=11)
ax.legend(loc="upper right", fontsize=9)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
ax.grid(True, alpha=0.3)
fig.tight_layout()
plot_zoom = PLOT_DIR / "surprise_forecast_zoom_w1.png"
fig.savefig(plot_zoom, dpi=120, bbox_inches="tight"); plt.close(fig)

# ── Plot 3: scatter (actual vs forecast) ─────────────────────────────
fig, ax = plt.subplots(figsize=(7, 7))
ax.scatter(y, p, s=4, alpha=0.4, color="tab:blue")
mx = max(y.max(), p.max()) * 1.05
ax.plot([0, mx], [0, mx], color="black", lw=1, ls="--", label="y = ŷ")
ax.set_xlabel("Actual load (kW)"); ax.set_ylabel("Forecast (kW)")
ax.set_title("Forecast vs Actual — March 2026", fontsize=11)
ax.set_xlim(0, mx); ax.set_ylim(0, mx)
ax.grid(True, alpha=0.3); ax.legend()
ax.set_aspect("equal")
fig.tight_layout()
plot_scatter = PLOT_DIR / "surprise_forecast_scatter.png"
fig.savefig(plot_scatter, dpi=120, bbox_inches="tight"); plt.close(fig)

# ── Plot 4: residual distribution ────────────────────────────────────
res = y - p
fig, ax = plt.subplots(figsize=(8, 4))
ax.hist(res, bins=80, alpha=0.7, color="tab:blue")
ax.axvline(0, color="black", lw=1)
ax.axvline(res.mean(), color="red", lw=1, ls="--", label=f"mean={res.mean():.3f}")
ax.set_xlabel("Residual (actual − forecast, kW)"); ax.set_ylabel("Count")
ax.set_title(f"Residual distribution — March 2026  |  std={res.std():.3f} kW", fontsize=11)
ax.grid(True, alpha=0.3); ax.legend()
fig.tight_layout()
plot_res = PLOT_DIR / "surprise_residual_hist.png"
fig.savefig(plot_res, dpi=120, bbox_inches="tight"); plt.close(fig)

print(f"Plots saved to {PLOT_DIR}")

# ── Build Excel ───────────────────────────────────────────────────────
out_xlsx = ROOT / "outputs/Surprise_Submission_Forecast.xlsx"
with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    # Sheet 1: Forecast (only timestamp + load_pred)
    forecast_out = forecast.rename(columns={"load_pred": "load_kw_predicted"})
    forecast_out.to_excel(writer, sheet_name="Forecast", index=False)

    # Sheet 2: Metrics
    m = metrics["metrics"]
    summary_df = pd.DataFrame([
        {"Metric": "RMSE",  "Value": round(m["rmse"], 4),  "Unit": "kW"},
        {"Metric": "MAE",   "Value": round(m["mae"], 4),   "Unit": "kW"},
        {"Metric": "MAPE",  "Value": round(m["mape"], 2),  "Unit": "%"},
        {"Metric": "sMAPE", "Value": round(m["smape"], 2), "Unit": "%"},
        {"Metric": "NRMSE", "Value": round(m["nrmse"], 2), "Unit": "% (★ primary ranking metric)"},
    ])
    summary_df.to_excel(writer, sheet_name="Metrics", index=False)
    info_df = pd.DataFrame([
        {"Field": "Method",     "Value": metrics["winner"]},
        {"Field": "Train range","Value": "2024-11-25 to 2026-02-28 (~16 months)"},
        {"Field": "Test range", "Value": "2026-03-01 to 2026-03-31 (last month)"},
        {"Field": "Train rows", "Value": 44256},
        {"Field": "Test rows",  "Value": 2976},
        {"Field": "Granularity","Value": "15 min"},
        {"Field": "Pipeline",   "Value": "NNLS blend (online_retraining + LSTM-AE + 8-bag LGBM) + MA(3) smoothing + variance-preserving alpha rescale"},
    ])
    info_df.to_excel(writer, sheet_name="Metrics", index=False, startrow=8)

    # Sheet 3: Full comparison table of all candidates
    cand_rows = []
    for c in metrics["all_candidates"]:
        cand_rows.append({"Method": c["name"], "RMSE (kW)": round(c["rmse"], 4),
                          "MAE (kW)": round(c["mae"], 4), "MAPE (%)": round(c["mape"], 2),
                          "sMAPE (%)": round(c["smape"], 2), "NRMSE (%)": round(c["nrmse"], 2)})
    cand_df = pd.DataFrame(sorted(cand_rows, key=lambda r: r["NRMSE (%)"]))
    cand_df.to_excel(writer, sheet_name="Comparison", index=False)

    # Sheet 4: Plots (placeholders — added after writer closes)

# Embed images using openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

wb = load_workbook(out_xlsx)
plots_ws = wb.create_sheet("Plots")
plots_ws["A1"] = "Forecast plots — March 2026 surprise dataset"
plots_ws["A1"].font = wb["Forecast"]["A1"].font.copy(bold=True, size=14)

current_row = 3
for img_path in [plot_full, plot_zoom, plot_scatter, plot_res]:
    img = XLImage(str(img_path))
    # Resize to ~ 1100 px wide
    if img.width > 1100:
        ratio = 1100 / img.width
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
    plots_ws.add_image(img, f"A{current_row}")
    rows_used = int(img.height / 18) + 3
    current_row += rows_used

wb.save(out_xlsx)
print(f"Saved Excel -> {out_xlsx}")

# Also save the timestamp+pred CSV separately as a clean submission file
csv_out = ROOT / "outputs/Surprise_Submission_Forecast.csv"
forecast.rename(columns={"load_pred": "load_kw_predicted"}).to_csv(csv_out, index=False)
print(f"Saved CSV   -> {csv_out}")
