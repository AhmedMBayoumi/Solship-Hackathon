"""
Build the official submission Excel file per supervisor's spec:
  - Forecasted vs actual load (data columns)
  - All errors calculated: MAE, MAPE, RMSE, NRMSE
  - Embedded plots: load vs forecast graph + error histograms
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parents[1]))

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import io

# Need openpyxl for image embedding
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage

ROOT = Path(__file__).parents[1]
SUB  = ROOT / "submission"
SUB_PLOTS = SUB / "plots"

df = pd.read_csv(ROOT / "data/processed/dataset_processed.csv", parse_dates=["timestamp"])
df_2025 = df[df["timestamp"].dt.year == 2025].copy()
preds = pd.read_csv(ROOT / "outputs/forecasts/bagging_walkforward_test_preds.csv", parse_dates=["timestamp"])

merged = df_2025[df_2025["timestamp"].dt.month.isin([4, 9])][["timestamp","load_kw"]].merge(
    preds, on="timestamp", how="left"
)
merged["error"]     = merged["load_kw"] - merged["load_pred"]
merged["abs_error"] = merged["error"].abs()
merged["pct_error"] = np.where(merged["load_kw"] > 0.01,
                                merged["abs_error"] / merged["load_kw"] * 100, 0)

def metrics(y, yp):
    rmse = float(np.sqrt(np.mean((y - yp) ** 2)))
    mae  = float(np.mean(np.abs(y - yp)))
    nrm  = rmse / np.mean(y) * 100
    # MAPE (use safe denom > 0.05 to avoid blow-up)
    mask = y > 0.05
    mape = float(np.mean(np.abs((y[mask] - yp[mask]) / y[mask])) * 100)
    r2   = 1 - np.sum((y - yp) ** 2) / np.sum((y - np.mean(y)) ** 2)
    mbe  = float(np.mean(yp - y))
    return {"RMSE_kW":round(rmse,4), "MAE_kW":round(mae,4), "MAPE_%":round(mape,2),
            "NRMSE_%":round(nrm,2), "R2":round(r2,4), "MBE_kW":round(mbe,4),
            "n":len(y), "mean_actual_kW":round(float(np.mean(y)),4)}

# Generate per-month plots (load+forecast, error histogram, scatter)
def make_month_plots(month, name, df_m):
    err = df_m["error"].values

    # 1. Time series
    fig, axes = plt.subplots(2, 1, figsize=(13, 6), sharex=True)
    axes[0].plot(df_m["timestamp"], df_m["load_kw"], color="#0066cc", lw=0.8, label="Actual")
    axes[0].plot(df_m["timestamp"], df_m["load_pred"], color="#cc3333", lw=0.8, ls="--", label="Forecast")
    axes[0].set_ylabel("Load (kW)")
    axes[0].set_title(f"{name} 2025 - Load forecast vs actual (walkforward bagging, 12x LightGBM)")
    axes[0].grid(alpha=0.3); axes[0].legend()
    axes[1].plot(df_m["timestamp"], err, color="#9933cc", lw=0.8)
    axes[1].axhline(0, color="black", lw=0.5)
    axes[1].fill_between(df_m["timestamp"], 0, err, where=(err > 0), alpha=0.25, color="green", label="under-pred")
    axes[1].fill_between(df_m["timestamp"], 0, err, where=(err < 0), alpha=0.25, color="red",   label="over-pred")
    axes[1].set_ylabel("Residual (kW)")
    axes[1].grid(alpha=0.3); axes[1].legend()
    axes[1].xaxis.set_major_locator(mdates.DayLocator(interval=2))
    axes[1].xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
    plt.setp(axes[1].xaxis.get_majorticklabels(), rotation=45)
    plt.tight_layout()
    p_ts = SUB_PLOTS / f"forecast_{name.lower()}_2025.png"
    plt.savefig(p_ts, dpi=130, bbox_inches="tight"); plt.close()

    # 2. Error histogram
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5))
    axes[0].hist(err, bins=60, color="#0066cc", alpha=0.75, edgecolor="black")
    axes[0].axvline(0, color="red", ls="--", lw=1.5, label="zero error")
    axes[0].axvline(err.mean(), color="green", lw=1.5, label=f"mean = {err.mean():+.4f}")
    axes[0].set_xlabel("Residual = actual - predicted (kW)")
    axes[0].set_ylabel("Count")
    axes[0].set_title(f"{name} 2025 - Residual distribution"); axes[0].legend(); axes[0].grid(alpha=0.3)
    abs_err = np.abs(err)
    axes[1].hist(abs_err, bins=60, color="#cc6600", alpha=0.75, edgecolor="black")
    axes[1].axvline(abs_err.mean(),   color="green",  lw=1.5, label=f"MAE = {abs_err.mean():.4f}")
    axes[1].axvline(np.median(abs_err), color="purple", lw=1.5, ls=":", label=f"median = {np.median(abs_err):.4f}")
    axes[1].set_xlabel("Absolute error |actual - predicted| (kW)")
    axes[1].set_title(f"{name} 2025 - Absolute error distribution"); axes[1].legend(); axes[1].grid(alpha=0.3)
    plt.tight_layout()
    p_h = SUB_PLOTS / f"error_hist_{name.lower()}_2025.png"
    plt.savefig(p_h, dpi=130, bbox_inches="tight"); plt.close()

    # 3. Scatter
    fig, ax = plt.subplots(figsize=(6.5, 6.5))
    ax.scatter(df_m["load_pred"], df_m["load_kw"], s=8, alpha=0.3, color="#0066cc")
    mx = max(df_m["load_kw"].max(), df_m["load_pred"].max()) * 1.05
    ax.plot([0,mx],[0,mx], "r--", lw=1.5, label="y = y-hat")
    ax.set_xlabel("Predicted (kW)"); ax.set_ylabel("Actual (kW)")
    ax.set_title(f"{name} 2025 - Pred vs actual")
    ax.set_xlim(0, mx); ax.set_ylim(0, mx); ax.grid(alpha=0.3); ax.legend()
    plt.tight_layout()
    p_sc = SUB_PLOTS / f"scatter_{name.lower()}_2025.png"
    plt.savefig(p_sc, dpi=130, bbox_inches="tight"); plt.close()
    return p_ts, p_h, p_sc

apr = merged[merged["timestamp"].dt.month == 4].copy()
sep = merged[merged["timestamp"].dt.month == 9].copy()
apr_ts, apr_h, apr_sc = make_month_plots(4, "April",     apr)
sep_ts, sep_h, sep_sc = make_month_plots(9, "September", sep)

apr_m = metrics(apr["load_kw"].values, apr["load_pred"].values)
sep_m = metrics(sep["load_kw"].values, sep["load_pred"].values)
all_m = metrics(merged["load_kw"].values, merged["load_pred"].values)

# ── Build the Excel file ──────────────────────────────────
wb = Workbook()
header_fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
metric_fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

def add_data_sheet(name, df_m, m_dict, ts_img, hist_img, sc_img):
    ws = wb.create_sheet(name)
    # Header / title
    ws["A1"] = f"{name} 2025 - Load forecast (walkforward bagging, 12x LightGBM)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    # Metrics block
    ws["A3"] = "Error metric"; ws["B3"] = "Value"
    for c in ["A3","B3"]:
        ws[c].fill = header_fill; ws[c].font = header_font
    metric_rows = [
        ("RMSE (kW)",   m_dict["RMSE_kW"]),
        ("MAE  (kW)",   m_dict["MAE_kW"]),
        ("MAPE (%)",    m_dict["MAPE_%"]),
        ("NRMSE (%)",   m_dict["NRMSE_%"]),
        ("R2",          m_dict["R2"]),
        ("Mean bias error MBE (kW)", m_dict["MBE_kW"]),
        ("N samples",   m_dict["n"]),
        ("Mean actual load (kW)", m_dict["mean_actual_kW"]),
    ]
    for i, (k, v) in enumerate(metric_rows):
        r = 4 + i
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        ws[f"A{r}"].fill = metric_fill

    # Data table
    data = df_m[["timestamp","load_kw","load_pred","error","abs_error","pct_error"]].copy()
    data.columns = ["timestamp","load_actual_kW","load_forecast_kW","error_kW","abs_error_kW","pct_error_%"]
    start = 14
    ws.cell(row=start, column=1).value = "DATA"
    ws.cell(row=start, column=1).font = Font(bold=True)
    for j, h in enumerate(data.columns, start=1):
        c = ws.cell(row=start+1, column=j); c.value = h; c.fill = header_fill; c.font = header_font
    for i, row in enumerate(data.itertuples(index=False), start=start+2):
        ws.cell(row=i, column=1).value = row.timestamp.strftime("%Y-%m-%d %H:%M")
        for j, v in enumerate(row[1:], start=2):
            ws.cell(row=i, column=j).value = round(float(v), 4) if v is not None else None

    # Column widths
    for col, w in zip("ABCDEF", [22, 18, 20, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    # Embed plots — anchor in column G
    img1 = XLImage(str(ts_img));  img1.width  = 850; img1.height = 380
    ws.add_image(img1, "G3")
    img2 = XLImage(str(hist_img)); img2.width = 850; img2.height = 320
    ws.add_image(img2, "G24")
    img3 = XLImage(str(sc_img));  img3.width  = 420; img3.height = 420
    ws.add_image(img3, "G42")

# Remove default sheet
default = wb.active; wb.remove(default)

# Summary sheet
ws = wb.create_sheet("Summary")
ws["A1"] = "Solship Energy AI Hackathon 2026 - Submission summary"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:F1")
ws["A2"] = "Best model: walkforward bagging (12x LightGBM, trained 2024 + 2025 up-to-test-month per supervisor)"
ws["A3"] = "MPC horizon: H = 96 (1 day, per supervisor's recommendation)"
ws["A4"] = "Test window: April 2025 + September 2025"
ws["A5"] = "Site: Sondrio, Italy   |   Battery 16 kWh / +-8 kW   |   Grid +-6 kW"

ws["A7"] = "FORECAST QUALITY"; ws["A7"].font = Font(bold=True, size=12)
hdr = ["month","RMSE_kW","MAE_kW","MAPE_%","NRMSE_%","R2","MBE_kW","n","mean_actual_kW"]
for j, h in enumerate(hdr, start=1):
    c = ws.cell(row=8, column=j); c.value = h; c.fill = header_fill; c.font = header_font
for i, (mon, m) in enumerate([("April", apr_m), ("September", sep_m), ("Combined", all_m)], start=9):
    ws.cell(row=i, column=1).value = mon
    for j, k in enumerate(hdr[1:], start=2):
        ws.cell(row=i, column=j).value = m.get(k, m.get(k.replace("_%","")))

# Bills
from src.eval.compute_bill import baseline_a_bill, baseline_b_bill, compute_bill
def per_month_bill(bf, df_):
    return {n: bf(df_[df_["timestamp"].dt.month == m]) for m, n in [(4,"April"),(9,"September")]}
A_per = per_month_bill(baseline_a_bill, df_2025)
B_per = per_month_bill(baseline_b_bill, df_2025)
mpc_df = pd.read_parquet(ROOT / "outputs/mpc_walkforward_H96.parquet")
mpc_per = {n: compute_bill(mpc_df[mpc_df["timestamp"].dt.month == m], mpc_df[mpc_df["timestamp"].dt.month == m]["p_grid_kw"]) for m, n in [(4,"April"),(9,"September")]}
oracle_per = {"April":     {"net_bill": -20.13},
              "September": {"net_bill":  -0.02}}

ws["A14"] = "CONTROLLER BILLS (EUR)"; ws["A14"].font = Font(bold=True, size=12)
hdr2 = ["controller","April_EUR","September_EUR","Total_EUR","Savings_vs_A_EUR","Savings_vs_A_pct"]
for j, h in enumerate(hdr2, start=1):
    c = ws.cell(row=15, column=j); c.value = h; c.fill = header_fill; c.font = header_font

A_total = A_per["April"]["net_bill"] + A_per["September"]["net_bill"]
controllers = [
    ("Baseline A (existing)", A_per["April"]["net_bill"], A_per["September"]["net_bill"]),
    ("Baseline B (no battery)", B_per["April"]["net_bill"], B_per["September"]["net_bill"]),
    ("Our MPC (walkforward bagging, H=96)", mpc_per["April"]["net_bill"], mpc_per["September"]["net_bill"]),
    ("Oracle (perfect foresight)", oracle_per["April"]["net_bill"], oracle_per["September"]["net_bill"]),
]
for i, (name, a, s) in enumerate(controllers, start=16):
    tot = a + s; sav = A_total - tot; pct = sav / abs(A_total) * 100
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=2).value = round(a, 2)
    ws.cell(row=i, column=3).value = round(s, 2)
    ws.cell(row=i, column=4).value = round(tot, 2)
    ws.cell(row=i, column=5).value = round(sav, 2)
    ws.cell(row=i, column=6).value = round(pct, 1)

for col, w in zip("ABCDEFGHI", [40,12,14,12,14,8,12,8,16]):
    ws.column_dimensions[col].width = w

# Embed bills bar chart and overall plots
ws.add_image(XLImage(str(SUB_PLOTS / "bills_comparison.png")).__class__(str(SUB_PLOTS / "bills_comparison.png")), "A22")

# April + September data sheets
add_data_sheet("April",     apr, apr_m, apr_ts, apr_h, apr_sc)
add_data_sheet("September", sep, sep_m, sep_ts, sep_h, sep_sc)

xlsx_path = SUB / "data" / "Solship_Hackathon_Submission.xlsx"
wb.save(xlsx_path)
print(f"Saved -> {xlsx_path}")
print(f"\nForecast metrics:")
for mon, m in [("April", apr_m), ("September", sep_m), ("Combined", all_m)]:
    print(f"  {mon:<10s} RMSE={m['RMSE_kW']:.4f}  MAE={m['MAE_kW']:.4f}  MAPE={m['MAPE_%']:.2f}%  NRMSE={m['NRMSE_%']:.2f}%  R2={m['R2']:.3f}")
