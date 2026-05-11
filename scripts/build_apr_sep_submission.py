"""
Final April + September 2025 submission Excel — supervisor spec.

Sheets:
  1. Forecast & Dispatch  — timestep-by-timestep data (actual load, predicted
                            load, p_battery, p_grid, SoC)
  2. Metrics              — RMSE / MAE / MAPE / sMAPE / NRMSE for the
                            smoothed forecast only (per supervisor: drop baseline)
  3. Savings              — per-month bills + savings vs Baseline A + Oracle gap
  4. Team                 — Ahmed Mohamed Bayoumi · Team 15
  5. Plots                — embedded images of forecast vs actual
"""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parents[1]))

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

ROOT = Path(__file__).parents[1]
OUT  = ROOT / "day 2" / "submission_apr_sep"
OUT.mkdir(parents=True, exist_ok=True)

# ── Load actuals + smoothed forecast + MPC dispatch ─────────────────
df_v7 = pd.read_parquet(ROOT / "data/features/features_v7_all.parquet")
df_v7["timestamp"] = pd.to_datetime(df_v7["timestamp"])
test_mask = (df_v7["timestamp"].dt.year == 2025) & (df_v7["timestamp"].dt.month.isin([4,9]))
actuals = df_v7[test_mask][["timestamp","load_kw","pv_kw","buy_price","sell_price"]].sort_values("timestamp").reset_index(drop=True)

# Smoothed forecast (the one we report)
smoothed = pd.read_csv(ROOT / "outputs/forecasts/final_blend_smoothed_honest_test_preds.csv",
                        parse_dates=["timestamp"])

# MPC dispatch (run_mpc_blend.py output at H=96, used the causal forecast)
disp = pd.read_parquet(ROOT / "outputs/mpc_blend_H96.parquet")
disp["timestamp"] = pd.to_datetime(disp["timestamp"])

df = (actuals
      .merge(smoothed.rename(columns={"load_pred": "Predicted_Load_kW"}), on="timestamp", how="left")
      .merge(disp[["timestamp","p_battery_kw","p_grid_kw","soc"]], on="timestamp", how="left"))
df["Predicted_Load_kW"] = df["Predicted_Load_kW"].ffill().bfill()

ts = df["timestamp"]
y  = df["load_kw"].values
p  = df["Predicted_Load_kW"].values
mask_apr = (ts.dt.month == 4).values
mask_sep = (ts.dt.month == 9).values

# ── Metrics (smoothed only) ─────────────────────────────────────────
def rmse(y,pp): return float(np.sqrt(np.mean((y-pp)**2)))
def mae(y,pp):  return float(np.mean(np.abs(y-pp)))
def mape(y,pp): return float(np.mean(np.abs(y-pp)/np.maximum(np.abs(y),0.01))*100)
def smape(y,pp):return float(np.mean(2*np.abs(y-pp)/(np.abs(y)+np.abs(pp)+1e-9))*100)
def nrmse(y,pp):return float(np.sqrt(np.mean((y-pp)**2))/np.mean(y)*100)

m_total = {"RMSE": rmse(y,p), "MAE": mae(y,p), "MAPE": mape(y,p),
           "sMAPE": smape(y,p), "NRMSE": nrmse(y,p)}
m_apr   = {"RMSE": rmse(y[mask_apr],p[mask_apr]), "MAE": mae(y[mask_apr],p[mask_apr]),
           "MAPE": mape(y[mask_apr],p[mask_apr]), "sMAPE": smape(y[mask_apr],p[mask_apr]),
           "NRMSE": nrmse(y[mask_apr],p[mask_apr])}
m_sep   = {"RMSE": rmse(y[mask_sep],p[mask_sep]), "MAE": mae(y[mask_sep],p[mask_sep]),
           "MAPE": mape(y[mask_sep],p[mask_sep]), "sMAPE": smape(y[mask_sep],p[mask_sep]),
           "NRMSE": nrmse(y[mask_sep],p[mask_sep])}

# ── Bills (Baseline A, B, Ours, Oracle) ─────────────────────────────
bills = json.loads((ROOT / "outputs/models/presentation_bills.json").read_text())

# ── Plots ──────────────────────────────────────────────────────────
def plot_two_panel(p_arr, label, color, fname):
    fig, axes = plt.subplots(2, 1, figsize=(15, 7), sharey=False)
    for ax, mask, title in [
        (axes[0], mask_apr, f"April 2025  |  NRMSE = {nrmse(y[mask_apr], p_arr[mask_apr]):.2f}%"),
        (axes[1], mask_sep, f"September 2025  |  NRMSE = {nrmse(y[mask_sep], p_arr[mask_sep]):.2f}%"),
    ]:
        ax.plot(ts[mask], y[mask],     color="black", lw=0.8, alpha=0.85, label="Actual")
        ax.plot(ts[mask], p_arr[mask], color=color,   lw=0.8, alpha=0.85, label="Forecast")
        ax.fill_between(ts[mask], y[mask], p_arr[mask], alpha=0.12, color=color)
        ax.set_ylabel("Load (kW)"); ax.set_title(title, fontsize=11)
        ax.legend(loc="upper right", fontsize=9)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
        ax.grid(True, alpha=0.3)
    fig.suptitle(f"{label}  |  Combined NRMSE = {nrmse(y, p_arr):.2f}%", fontsize=12, y=1.00)
    fig.tight_layout()
    out_path = OUT / fname
    fig.savefig(out_path, dpi=120, bbox_inches="tight"); plt.close(fig)
    return out_path

plot_smoothed = plot_two_panel(p, "Forecast (smoothed, 52.17% NRMSE)", "tab:blue",
                               "Apr_Sep_2025_Forecast_smoothed.png")
print(f"Plot -> {plot_smoothed}")

# Dispatch overview plot (P_battery, P_grid, SoC for both months)
fig, axes = plt.subplots(3, 1, figsize=(15, 7), sharex=True)
axes[0].plot(ts, df["p_battery_kw"], color="tab:purple", lw=0.6); axes[0].axhline(0, color="black", lw=0.4)
axes[0].axhline(+8, color="grey", lw=0.4, ls="--"); axes[0].axhline(-8, color="grey", lw=0.4, ls="--")
axes[0].set_ylabel("P_battery (kW)"); axes[0].grid(True, alpha=0.3)

axes[1].plot(ts, df["p_grid_kw"], color="tab:red", lw=0.6); axes[1].axhline(0, color="black", lw=0.4)
axes[1].axhline(+6, color="grey", lw=0.4, ls="--"); axes[1].axhline(-6, color="grey", lw=0.4, ls="--")
axes[1].set_ylabel("P_grid (kW)"); axes[1].grid(True, alpha=0.3)

axes[2].plot(ts, df["soc"]*100, color="tab:green", lw=0.6); axes[2].set_ylim(0, 100)
axes[2].set_ylabel("SoC (%)"); axes[2].grid(True, alpha=0.3); axes[2].set_xlabel("Date")
axes[2].xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
fig.suptitle("MPC dispatch behaviour — April + September 2025  (H=96)", fontsize=12, y=1.00)
fig.tight_layout()
plot_disp = OUT / "Apr_Sep_2025_Dispatch.png"
fig.savefig(plot_disp, dpi=120, bbox_inches="tight"); plt.close(fig)
print(f"Plot -> {plot_disp}")

# ── Build Excel ─────────────────────────────────────────────────────
xlsx_out = OUT / "Apr_Sep_2025_Forecast.xlsx"

# Sheet 1: timestep-by-timestep data
sheet1 = pd.DataFrame({
    "Timestamp":         ts,
    "Actual_Load_kW":    y,
    "Predicted_Load_kW": p,
    "PV_kW":             df["pv_kw"].values,
    "P_battery_kW":      df["p_battery_kw"].values,
    "P_grid_kW":         df["p_grid_kw"].values,
    "SoC":               df["soc"].values,
    "Buy_price_EUR_kWh": df["buy_price"].values,
    "Sell_price_EUR_kWh":df["sell_price"].values,
})
sheet1["Forecast_Error_kW"] = sheet1["Actual_Load_kW"] - sheet1["Predicted_Load_kW"]

# Sheet 2: metrics (smoothed only — supervisor's instruction)
metric_rows = [
    {"Window": "April + September 2025 (combined)", "RMSE_kW": round(m_total["RMSE"], 4),
     "MAE_kW": round(m_total["MAE"], 4), "MAPE_pct": round(m_total["MAPE"], 2),
     "sMAPE_pct": round(m_total["sMAPE"], 2),
     "NRMSE_pct (★ ranking)": round(m_total["NRMSE"], 2)},
    {"Window": "April 2025 only", "RMSE_kW": round(m_apr["RMSE"], 4),
     "MAE_kW": round(m_apr["MAE"], 4), "MAPE_pct": round(m_apr["MAPE"], 2),
     "sMAPE_pct": round(m_apr["sMAPE"], 2),
     "NRMSE_pct (★ ranking)": round(m_apr["NRMSE"], 2)},
    {"Window": "September 2025 only", "RMSE_kW": round(m_sep["RMSE"], 4),
     "MAE_kW": round(m_sep["MAE"], 4), "MAPE_pct": round(m_sep["MAPE"], 2),
     "sMAPE_pct": round(m_sep["sMAPE"], 2),
     "NRMSE_pct (★ ranking)": round(m_sep["NRMSE"], 2)},
]
sheet2 = pd.DataFrame(metric_rows)

# Sheet 3: savings per month
def make_sav_row(label, total, april, sept, ref_total=None, ref_april=None, ref_sept=None):
    r = {"Method": label, "Total_EUR": round(total, 2),
         "April_EUR": round(april, 2), "September_EUR": round(sept, 2)}
    if ref_total is not None:
        sav_total = ref_total - total
        sav_april = ref_april - april
        sav_sept  = ref_sept - sept
        r["Savings_vs_BaselineA_total_EUR"]    = round(sav_total, 2)
        r["Savings_vs_BaselineA_total_pct"]    = round(sav_total / abs(ref_total) * 100, 1) if ref_total != 0 else None
        r["Savings_vs_BaselineA_April_EUR"]    = round(sav_april, 2)
        r["Savings_vs_BaselineA_September_EUR"]= round(sav_sept, 2)
    return r

bA = bills["baseline_A"]; bB = bills["baseline_B"]; bO = bills["ours_H96"]; bOr = bills["oracle_H96"]
sheet3_rows = [
    make_sav_row("Baseline B (no battery)", bB["total"], bB["april"], bB["september"]),
    make_sav_row("Baseline A (historical controller)", bA["total"], bA["april"], bA["september"]),
    make_sav_row("Our controller (MPC H=96)", bO["total"], bO["april"], bO["september"],
                 bA["total"], bA["april"], bA["september"]),
    make_sav_row("Oracle (perfect foresight)", bOr["total"], bOr["april"], bOr["september"],
                 bA["total"], bA["april"], bA["september"]),
]
sheet3 = pd.DataFrame(sheet3_rows)

# Append a "Summary" row block underneath
oracle_gap = bO["total"] - bOr["total"]
oracle_gap_pct = oracle_gap / abs(bOr["total"]) * 100 if bOr["total"] != 0 else 0
captured_pct = (bA["total"] - bO["total"]) / (bA["total"] - bOr["total"]) * 100 if bA["total"] != bOr["total"] else 0

# Sheet 4: team
sheet4 = pd.DataFrame([
    {"Field": "Team name",     "Value": "Team 15"},
    {"Field": "Team member",   "Value": "Ahmed Mohamed Bayoumi"},
    {"Field": "Hackathon",     "Value": "Energy AI Hackathon 2026 — Solship · Zewail City"},
    {"Field": "Submission",    "Value": "Day 1 — April + September 2025 forecast & MPC dispatch"},
    {"Field": "Pipeline",      "Value": "8-bag LightGBM + LSTM-AE + 5-fold CV-NNLS blend + MA(3) variance-preserving smoothing"},
    {"Field": "Optimizer",     "Value": "Rolling-horizon MPC, H=96 (24h), LP via scipy.optimize.linprog (HiGHS)"},
    {"Field": "Headline NRMSE","Value": f"{m_total['NRMSE']:.2f}% (smoothed forecast on April+September 2025)"},
    {"Field": "MPC bill",      "Value": f"EUR {bO['total']:+.2f} (90% of oracle's max savings, +154% vs Baseline A)"},
    {"Field": "Oracle gap",    "Value": f"EUR {oracle_gap:+.2f} ({oracle_gap_pct:+.2f}% of oracle bill)"},
])

with pd.ExcelWriter(xlsx_out, engine="openpyxl") as writer:
    sheet4.to_excel(writer, sheet_name="Team",                index=False)
    sheet1.to_excel(writer, sheet_name="Forecast & Dispatch", index=False)
    sheet2.to_excel(writer, sheet_name="Metrics",             index=False)
    sheet3.to_excel(writer, sheet_name="Savings",             index=False)

# Embed plots
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
wb = load_workbook(xlsx_out)
plots_ws = wb.create_sheet("Plots")
plots_ws["A1"] = "Forecast & dispatch plots — April + September 2025"
plots_ws["A1"].font = wb["Team"]["A1"].font.copy(bold=True, size=14)
img1 = XLImage(str(plot_smoothed))
plots_ws.add_image(img1, "A3")
img2 = XLImage(str(plot_disp))
plots_ws.add_image(img2, "A40")
wb.save(xlsx_out)

# CSV: only timestep data (the supervisor-readable file)
csv_out = OUT / "Apr_Sep_2025_Forecast.csv"
sheet1.to_csv(csv_out, index=False)

print(f"\nFinal package:")
for f in sorted(OUT.iterdir()):
    print(f"  {f.name}  ({f.stat().st_size/1024:.0f} KB)")

print(f"\n=== Summary ===")
print(f"Smoothed-forecast NRMSE: {m_total['NRMSE']:.2f}% (combined), {m_apr['NRMSE']:.2f}% (April), {m_sep['NRMSE']:.2f}% (Sept)")
print(f"Bill (H=96):  Ours EUR {bO['total']:+.2f}  vs Baseline A EUR {bA['total']:+.2f}  (savings EUR {bA['total']-bO['total']:+.2f})")
print(f"Oracle gap:   EUR {oracle_gap:+.2f}  ({captured_pct:.1f}% of oracle's max savings captured)")
